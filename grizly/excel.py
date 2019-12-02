import os
import pandas as pd
import openpyxl

def read_excel(excel_path, sheet_name="", query=""):
    if sheet_name != "":
        fields = pd.read_excel(excel_path, sheet_name=sheet_name).fillna("")
    else:
        fields = pd.read_excel(excel_path).fillna("")

    schema = "" if "schema" not in fields else fields["schema"][0]
    table = fields["table"][0]

    if query != "":
        fields = fields.query(query)

    columns_qf = {}
    for index, row in fields.iterrows():
        if row["column"] == "":
            attr = row["column_as"]
        else:
            attr = row["column"]
        columns_qf[attr] = {}
        columns_qf[attr]["type"] = row["column_type"]
        columns_qf[attr]["group_by"] = row["group_by"]
        try:
            if row["expression"] != "":
                columns_qf[attr]["expression"] = row["expression"]
        except:
            pass
        if row["column_as"] != "":
            columns_qf[attr]["as"] = row["column_as"]
        try:
            if row["select"] != "":
                columns_qf[attr]["select"] = row["select"]
        except:
            pass
        try:
            if row["custom_type"] != "":
                columns_qf[attr]["custom_type"] = row["custom_type"]
        except:
            pass

    return schema, table, columns_qf


def copy_df_to_excel(df, input_excel_path, output_excel_path, sheet_name='', startrow=0, startcol=0, index=False, header=False):
    writer = pd.ExcelWriter(input_excel_path, engine='openpyxl')
    book = openpyxl.load_workbook(input_excel_path)
    writer.book = book

    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    df.to_excel(writer, sheet_name=sheet_name,startrow=startrow,startcol=startcol,index=index,header=header)

    writer.path = output_excel_path
    writer.save()
    writer.close()


class Excel:
    """Class which deals with Excel files.
    """
    def __init__(self, excel_path, output_excel_path=''):
        """
        Parameters
        ----------
        excel_path : str
            Path to input Excel file.
        output_excel_path : str, optional
            Path to output Excel.
        """
        self.input_excel_path = excel_path
        self.filename = os.path.basename(self.input_excel_path)
        if output_excel_path != '':
            self.output_excel_path = output_excel_path
        else:
            self.output_excel_path = os.path.join(os.path.split(excel_path)[0], 
                                        os.path.splitext(self.filename)[0] + '_working' + os.path.splitext(self.filename)[1])
        #self.book = book = openpyxl.load_workbook(self.input_excel_path)
    

    def write_df(self, df, sheet, row=1, col=1, index=False, header=False):
        """Saves DatFrame in Excel file.
        
        Parameters
        ----------
        df : pandas.pd.DataFrame
            pd.DataFrame to be saved in Excel
        sheet: str
            Name of sheet
        row : int, optional
            Upper left cell row to dump pd.DataFrame, by default 1
        col : int, optional
            Upper left cell column to dump pd.DataFrame, by default 1
        index : bool, optional
            Write row names (index), by default False
        header : bool, optional
            Write column names (header), by default False
        """

        writer = pd.ExcelWriter(self.input_excel_path, engine='openpyxl')
        book = openpyxl.load_workbook(self.input_excel_path)
        writer.book = book

        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        df.to_excel(writer, sheet_name=sheet,startrow=row-1,startcol=col-1,index=index,header=header)

        writer.path = self.output_excel_path
        writer.save()
        writer.close()

        self.input_excel_path = self.output_excel_path
        self.filename = os.path.basename(self.output_excel_path)

        return self
        

    def write_value(self, sheet, row, col, value):
        """Writes cell value to Excel file.
        
        Parameters
        ----------
        sheet : str
            Name of sheet
        row : int
            Cell row
        col : int
            Cell column

        Returns
        -------
        float
            Cell value
        """
        book = openpyxl.load_workbook(self.input_excel_path)

        worksheet = book.get_sheet_by_name(sheet)
        worksheet.cell(row=row, column=col, value=value)
        book.save(filename = self.output_excel_path)
        
        print("Written value {} in sheet {}".format(value, sheet))

        return self
     

    def save(self):
        """save to workbook"""
        #self.book.save()
        pass


    def get_value(self, sheet, row, col):
        """Extracts cell value from Excel file.
        
        Parameters
        ----------
        sheet : str
            Name of sheet
        row : int
            Cell row
        col : int
            Cell column

        Returns
        -------
        float
            Cell value
        """
        
        wb = openpyxl.load_workbook(self.input_excel_path, data_only=True)
        sh = wb[sheet]
        value = sh.cell(row, col).value
        return value