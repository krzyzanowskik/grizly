import boto3
import os
import csv
from pandas import (
    ExcelWriter
)
import openpyxl
import win32com.client as win32
from grizly.core.utils import (
    get_path,
    read_config,
    check_if_exists
)
from pandas import DataFrame
from sqlalchemy import create_engine
from sqlalchemy.pool import NullPool
from io import StringIO

grizly_config = read_config()
os.environ["HTTPS_PROXY"] = grizly_config["https"]

class Csv():
    def __init__(self, config=None, engine_str:str=None):
        if config == None:
            self.engine_str = engine_str
        else:
            self.engine_str = config.engine_str
        self.deletethis = ""

    def from_sql(self, table, csv_path, schema:str=None, chunksize:int=1000):
        engine = create_engine(self.engine_str, encoding='utf8', poolclass=NullPool)
        conn = engine.connect().connection
        cursor = conn.cursor()
        sql = f"""SELECT count(*) FROM {table};"""
        cursor.execute(sql)
        records_count = cursor.fetchone()[0]
        chunks = 0
        chunks_int = records_count//chunksize
        chunks_float = records_count/chunksize
        if chunks_float>chunks_int:
            chunks += chunks_int + 1
        if chunks_float<1:
            chunks += chunks_int + 1
        for chunk in range(chunks):
            offset = chunk * chunksize
            sql = f"""SELECT * FROM {table} LIMIT {chunksize} OFFSET {offset};"""
            cursor.execute(sql)
            rows = cursor.fetchall()
            if not rows:
                break
            if chunk == 0:
                os.remove(csv_path)
            with open(csv_path, 'a', newline='', encoding = 'utf-8') as csvfile:
                writer = csv.writer(csvfile, delimiter=',')
                writer.writerows(rows)
        cursor.close()
        conn.close()
        return self

    def from_sfdc(self):
        pass
    
    def from_github(self):
        pass

class Excel:
    """Class which deals with Excel files.
    """
    def __init__(self, excel_path, output_excel_path=''):
        """
        Parameters
        ----------
        excel_path : str
            Path to input Excel file.
        output_excel_path : str, optional
            Path to output Excel.
        """
        self.input_excel_path = excel_path
        self.filename = os.path.basename(self.input_excel_path)
        if output_excel_path != '':
            self.output_excel_path = output_excel_path
        else:
            self.output_excel_path = os.path.join(os.path.split(excel_path)[0], 
                                        os.path.splitext(self.filename)[0] + '_working' + os.path.splitext(self.filename)[1])
        #self.book = book = openpyxl.load_workbook(self.input_excel_path)
    

    def write_df(self, df, sheet, row=1, col=1, index=False, header=False):
        """Saves DatFrame in Excel file.
        
        Parameters
        ----------
        df : pandas.DataFrame
            DataFrame to be saved in Excel
        sheet: str
            Name of sheet
        row : int, optional
            Upper left cell row to dump DataFrame, by default 1
        col : int, optional
            Upper left cell column to dump DataFrame, by default 1
        index : bool, optional
            Write row names (index), by default False
        header : bool, optional
            Write column names (header), by default False
        """

        writer = ExcelWriter(self.input_excel_path, engine='openpyxl')
        book = openpyxl.load_workbook(self.input_excel_path)
        writer.book = book

        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        df.to_excel(writer, sheet_name=sheet,startrow=row-1,startcol=col-1,index=index,header=header)

        writer.path = self.output_excel_path
        writer.save()
        writer.close()

        self.input_excel_path = self.output_excel_path
        self.filename = os.path.basename(self.output_excel_path)

        return self
        

    def write_value(self, sheet, row, col, value):
        """Writes cell value to Excel file.
        
        Parameters
        ----------
        sheet : str
            Name of sheet
        row : int
            Cell row
        col : int
            Cell column

        Returns
        -------
        float
            Cell value
        """
        book = openpyxl.load_workbook(self.input_excel_path)

        worksheet = book.get_sheet_by_name(sheet)
        worksheet.cell(row=row, column=col, value=value)
        book.save(filename = self.output_excel_path)
        
        print("Written value {} in sheet {}".format(value, sheet))

        return self
     

    def save(self):
        """save to workbook"""
        #self.book.save()
        pass


    def get_value(self, sheet, row, col):
        """Extracts cell value from Excel file.
        
        Parameters
        ----------
        sheet : str
            Name of sheet
        row : int
            Cell row
        col : int
            Cell column

        Returns
        -------
        float
            Cell value
        """
        xlApp = win32.Dispatch('Excel.Application')
        wb = xlApp.Workbooks.Open(self.input_excel_path)
        ws = wb.Worksheets(sheet)
        value = ws.Cells(row,col).Value
        wb.Close()
        xlApp.Quit()
        
        return value


    def open(self, input=False):
        """[summary]
        
        Parameters
        ----------
        input : bool, optional
            [description], by default False
        
        Returns
        -------
        [type]
            [description]
        """

        if input == False:
            path = self.input_excel_path
        else:
            path = self.output_excel_path

        try:
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            try:
                xlwb = excel.Workbooks(path)
            except Exception as e:
                try:
                    xlwb = excel.Workbooks.Open(path)
                except Exception as e:
                    print(e)
                    xlwb = None
            ws = xlwb.Worksheets('blaaaa') 
            excel.Visible = True

        except Exception as e:
            print(e)

        finally:
            ws = None
            wb = None
            excel = None

        return self


class AttrDict(dict):
    def __init__(self, *args, **kwargs):
        super(AttrDict, self).__init__(*args, **kwargs)
        self.__dict__ = self


class AWS:
    """Class that represents a file in S3.
    """
    def __init__(self, file_name:str=None, s3_key:str=None, bucket:str=None, file_dir:str=None, config=None):
        """
        Examples
        --------
        >>> from grizly import get_path, AWS
        >>> s3 = AWS('emea_daily.xlsx', s3_key='dbb/ENG_EMEA/', file_dir=get_path('acoe_projects', 'dbb', 'ENG_EMEA'))

        Parameters
        ----------
        file_name : str, optional
            Name of the file, if None then 'test.csv'
        s3_key : str, optional
            Name of s3 key, if None then 'bulk/'
        bucket : str, optional
            Bucket name, if None then 'teis-data'
        file_dir : str, optional
            Path to local folder to store the file, if None then 'C:\\Users\\your_id\\s3_loads'
        config : module, optional
            Config module (imported .py file), by default None
        """
        if not config:
            config = AttrDict()
            config.update({
                        'file_name': 'test.csv', 
                        's3_key' : 'bulk/',
                        'bucket' : 'teis-data',
                        'file_dir' : get_path('s3_loads')
                        })

        self.file_name = file_name if file_name else config.file_name
        self.s3_key = s3_key if s3_key else config.s3_key
        self.bucket = bucket if bucket else config.bucket
        self.file_dir = file_dir if file_dir else config.file_dir
        self.s3_resource = boto3.resource('s3', 
                            aws_access_key_id=grizly_config["akey"], 
                            aws_secret_access_key=grizly_config["skey"], 
                            region_name=grizly_config["region"])

        os.makedirs(self.file_dir, exist_ok=True)

    def info(self):
        """Print a concise summary of a AWS.

        Examples
        --------
        >>> AWS().info()
        """
        print(f"\033[1m file_name: \033[0m\t'{self.file_name}'")
        print(f"\033[1m s3_key: \033[0m\t'{self.s3_key}'")
        print(f"\033[1m bucket: \033[0m\t'{self.bucket}'")
        print(f"\033[1m file_dir: \033[0m\t'{self.file_dir}'")


    def file_to_s3(self):
        """Writes local file to s3.

        Examples
        --------
        >>> from grizly import get_path
        >>> file_dir=get_path('acoe_projects', 'analytics_project_starter', '01_workflows')
        >>> aws = AWS('test_table.csv', s3_key='analytics_project_starter/test/', file_dir=file_dir)
        >>> aws.file_to_s3()
        """
        file_path = os.path.join(self.file_dir, self.file_name)

        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File '{file_path}' does not exist.")

        s3_key = self.s3_key + self.file_name
        s3_file = self.s3_resource.Object(self.bucket, s3_key)
        s3_file.upload_file(file_path)

        print(f"'{self.file_name}' uploaded to '{self.bucket}' bucket as '{s3_key}'")


    def s3_to_file(self):
        """Writes s3 to local file.

        Examples
        --------
        >>> aws = AWS()
        >>> aws.s3_to_file()
        """
        file_path = os.path.join(self.file_dir, self.file_name)

        s3_key = self.s3_key + self.file_name
        s3_file = self.s3_resource.Object(self.bucket, s3_key)
        s3_file.download_file(file_path)

        print(f"'{s3_key}' uploaded to '{file_path}'")


    def df_to_s3(self, df:DataFrame):
        """Saves DataFrame in s3.
        
        Examples
        --------
        >>> from pandas import DataFrame
        >>> df = DataFrame({'col1': [1, 2], 'col2': [3, 4]})
        >>> AWS().df_to_s3(df)

        Parameters
        ----------
        df : DataFrame
            DataFrame object
        """
        if not isinstance(df, DataFrame):
            raise ValueError("'df' must be DataFrame object")

        file_path = os.path.join(self.file_dir, self.file_name)

        if not file_path.endswith('.csv'):
            raise ValueError("Invalid file extention - 'file_name' attribute must end with '.csv'")

        df.to_csv(file_path)
        print(f"DataFrame saved in '{file_path}'")

        self.file_to_s3()


    def s3_to_rds(self, table:str, schema:str=None, sep:str='\t', if_exists='fail'):
        """Writes S3 file to Redshift table.    

        Parameters
        ----------
        table : str
            Table name
        schema : str, optional
            Schame name, by default None
        if_exists : {'fail', 'replace', 'append'}, default 'fail'
            How to behave if the table already exists.

            * fail: Raise a ValueError.
            * replace: Clean table before inserting new values.
            * append: Insert new values to the existing table.

        sep : str, optional
            Separator, by default '\t'
        """
        if if_exists not in ("fail", "replace", "append"):
            raise ValueError(f"'{if_exists}' is not valid for if_exists")

        table_name = f'{schema}.{table}' if schema else table
        engine = create_engine("mssql+pyodbc://Redshift")

        if check_if_exists(table, schema):
            if if_exists == 'fail':
                raise ValueError("Table {} already exists".format(table_name))
            elif if_exists == 'replace':
                sql ="DELETE FROM {}".format(table_name)
                engine.execute(sql)
                print('SQL table has been cleaned up successfully.')
            else:
                pass
        else:
            self._create_table_like_s3(table_name, sep)

        s3_key = self.s3_key + self.file_name
        print("Loading {} data into {} ...".format(s3_key, table_name))

        sql = f"""
            COPY {table_name} FROM 's3://teis-data/{s3_key}'
            access_key_id '{grizly_config["akey"]}'
            secret_access_key '{grizly_config["skey"]}'
            delimiter '{sep}'
            NULL ''
            IGNOREHEADER 1
            REMOVEQUOTES
            ;commit;
            """

        engine.execute(sql)
        print(f'Data has been copied to {table_name}')


    def df_to_rds(self, df:DataFrame, table:str, schema:str=None, sep:str='\t'):
        """Writes DataFrame to Redshift.
        
        Parameters
        ----------
        df : DataFrame
            DataFrame object
        table : str
            Table name
        schema : str, optional
            Schema name, by default None
        sep : str, optional
            Separator, by default '\t'
        """
        self.df_to_s3(df)

        engine = create_engine("mssql+pyodbc://Redshift")

        if not check_if_exists(table, schema):
            df.head(1).to_sql(table, schema=schema, index=False, con=engine)
            table_name = f'{schema}.{table}' if schema else f'{table}'
            print(f'Table {table_name} has been created.')

        self.s3_to_rds(table, schema, sep=sep)


    def _create_table_like_s3(self, table_name, sep):
        s3_client = self.s3_resource.meta.client

        obj_content = s3_client.select_object_content(
                        Bucket=self.bucket,
                        Key=self.s3_key + self.file_name,
                        ExpressionType='SQL',
                        Expression="select * from s3object limit 10",
                        InputSerialization = {'CSV': {'FileHeaderInfo': 'None', 'FieldDelimiter': sep}},
                        OutputSerialization = {'CSV': {}},
                        )

        records = []
        for event in obj_content['Payload']:
            if 'Records' in event:
                records.append(event['Records']['Payload'])

        file_str = ''.join(r.decode('utf-8') for r in records)

        select_df = read_csv(StringIO(file_str))
        columns = []

        for col in select_df:
            if select_df[col].dtype == 'int64':
                columns.append(f"{col} INT")
            elif select_df[col].dtype == 'float':
                columns.append(f"{col} FLOAT")
            else:
                columns.append(f"{col} VARCHAR(500)")

        columns_str = ", ".join(columns)
        sql = "CREATE TABLE {} ({})".format(table_name, columns_str)

        engine = create_engine("mssql+pyodbc://Redshift")
        engine.execute(sql)

        print("Table {} has been created successfully.".format(table_name))